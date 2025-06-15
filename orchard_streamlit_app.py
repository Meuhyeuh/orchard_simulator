import streamlit as st
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

SUSCEPTIBILITY_OPTIONS = ["high", "medium", "low"]
CLUSTER_PENALTY = {
    "high": 0.1,
    "medium": 0.4,
    "low": 1.0
}

# === STREAMLIT UI ===
st.title("Apple Orchard Layout Simulator")
st.write("Design a multi-species orchard that limits disease spread.")

num_rows = st.slider("Number of Rows", 3, 20, 6)
num_cols = st.slider("Trees per Row", 10, 50, 25)

st.subheader("Define Your Tree Species")
default_species = pd.DataFrame([
    {"Species": "Golden", "Probability": 0.3, "Susceptibility": "high", "Color": "#7FB77E"},
    {"Species": "Liberty", "Probability": 0.4, "Susceptibility": "medium", "Color": "#F4A261"},
    {"Species": "Enterprise", "Probability": 0.3, "Susceptibility": "low", "Color": "#8D6E63"},
])

edited_species_df = st.data_editor(
    default_species,
    num_rows="dynamic",
    column_config={
        "Probability": st.column_config.NumberColumn(
            min_value=0.0, max_value=1.0, step=0.1, format="%.1f"
        ),
        "Susceptibility": st.column_config.SelectboxColumn(options=SUSCEPTIBILITY_OPTIONS),
        "Color": st.column_config.TextColumn()
    },
    use_container_width=True
)

# Enforce unique species names
if edited_species_df['Species'].duplicated().any():
    st.error("Species names must be unique.")
    st.stop()

# Enforce unique species names
if edited_species_df['Species'].duplicated().any():
    st.error("Species names must be unique.")
    st.stop()
        "Susceptibility": st.column_config.SelectboxColumn(options=SUSCEPTIBILITY_OPTIONS),
        "Color": st.column_config.TextColumn()  # Replace ColorColumn with TextColumn for compatibility
    },
    use_container_width=True
)

SPECIES = []
probabilities = {}
susceptibility_map = {}
COLOR_MAP = {}

# Optional: Allow user to assign custom colors in future version

for idx, (i, row) in enumerate(edited_species_df.iterrows()):
    species_id = idx
    name = row["Species"]
    prob = row["Probability"]
    susc = row["Susceptibility"]
    color = row["Color"]
    SPECIES.append({"id": species_id, "name": name})
    probabilities[species_id] = prob
    susceptibility_map[species_id] = susc
    COLOR_MAP[species_id] = color

ID_TO_NAME = {s["id"]: s["name"] for s in SPECIES}
ID_TO_COLOR = {s["id"]: COLOR_MAP[s["id"]] for s in SPECIES}

total = sum(probabilities.values())
if total == 0:
    st.error("Total probability cannot be zero.")
    st.stop()
elif abs(total - 1.0) > 0.01:
    st.warning(f"Total probability = {total:.2f}. It will be normalized automatically.")
norm_probs = {k: v / total for k, v in probabilities.items()}

def choose_species(i, j, grid):
    probs = []
    for s in SPECIES:
        prob = norm_probs[s["id"]]
        penalty = CLUSTER_PENALTY[susceptibility_map[s["id"]]]
        if penalty < 1.0:
            if j >= 2 and grid[i, j-1] == s["id"] and grid[i, j-2] == s["id"]:
                prob *= penalty
            if i >= 2 and grid[i-1, j] == s["id"] and grid[i-2, j] == s["id"]:
                prob *= penalty
        probs.append(prob)
    total = sum(probs)
    probs = [p / total for p in probs]
    return np.random.choice([s["id"] for s in SPECIES], p=probs)

def generate_orchard():
    grid = np.full((num_rows, num_cols), -1)
    for i in range(num_rows):
        for j in range(num_cols):
            grid[i, j] = choose_species(i, j, grid)
    return grid

if st.button("Simulate Orchard"):
    grid = generate_orchard()
    fig, ax = plt.subplots(figsize=(12, 2.5))
    color_array = np.vectorize(COLOR_MAP.get)(grid)
    ax.imshow(grid, cmap=None)
    ax.set_xticks([])
    ax.set_yticks([])

    for x in range(grid.shape[1] + 1):
        ax.axvline(x - 0.5, color='black', linewidth=0.5)
    for y in range(grid.shape[0] + 1):
        ax.axhline(y - 0.5, color='black', linewidth=0.5)

    handles = [plt.Rectangle((0,0),1,1, color=COLOR_MAP[s["id"]]) for s in SPECIES]
    labels = [s["name"] for s in SPECIES]
    ax.legend(handles, labels, loc='upper right', bbox_to_anchor=(1.15, 1))

    st.pyplot(fig)
    st.success("Simulation complete!")

    name_grid = np.vectorize(ID_TO_NAME.get)(grid)
    df = pd.DataFrame(name_grid)
    wb = Workbook()
    ws = wb.active
    ws.title = "Orchard Layout"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            species_id = next(s["id"] for s in SPECIES if s["name"] == val)
            hex_color = COLOR_MAP.get(species_id, "FFFFFF").replace("#", "")
            fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            cell.fill = fill

    for col in range(1, num_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 3
    for row in range(1, num_rows + 1):
        ws.row_dimensions[row].height = 20

    meta_ws = wb.create_sheet(title="Simulation Info")
    meta_ws.append(["Species", "Color", "Susceptibility", "Probability"])
    for s in SPECIES:
        meta_ws.append([
            s["name"],
            COLOR_MAP[s["id"]],
            susceptibility_map[s["id"]],
            round(probabilities[s["id"]], 2)
        ])
    meta_ws.append([])
    meta_ws.append(["Grid Size", f"{num_rows} rows x {num_cols} columns"])

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.download_button(
        label="Download Grid as Excel",
        data=excel_buffer,
        file_name="orchard_grid.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
